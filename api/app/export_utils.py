"""
Утилиты для экспорта результатов поиска
"""

from typing import List, Dict
import json
import csv
import io


class SearchResultsExporter:
    """Экспорт результатов поиска в различные форматы"""
    
    @staticmethod
    def to_json(results: List[Dict], pretty: bool = True) -> str:
        """Экспорт в JSON"""
        indent = 2 if pretty else None
        return json.dumps(results, ensure_ascii=False, indent=indent)
    
    @staticmethod
    def to_csv(results: List[Dict]) -> str:
        """Экспорт в CSV"""
        if not results:
            return ""
        
        output = io.StringIO()
        
        # Определяем поля по первой записи
        fieldnames = list(results[0].keys())
        
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        
        for result in results:
            # Убираем вложенные объекты для CSV
            flat_result = {}
            for key, value in result.items():
                if isinstance(value, (dict, list)):
                    flat_result[key] = json.dumps(value, ensure_ascii=False)
                else:
                    flat_result[key] = value
            writer.writerow(flat_result)
        
        return output.getvalue()
    
    @staticmethod
    def to_excel(results: List[Dict], filename: str = None):
        """
        Экспорт в Excel с форматированием:
        - Заголовки жирным
        - Автоширина колонок
        - Фильтры
        - Цветовое кодирование по релевантности
        """
        try:
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise ImportError("Для экспорта в Excel требуются pandas и openpyxl")
        
        if not results:
            return None
        
        # Преобразуем в DataFrame
        # Убираем вложенные объекты
        flat_results = []
        for result in results:
            flat_result = {}
            for key, value in result.items():
                if isinstance(value, (dict, list)):
                    flat_result[key] = json.dumps(value, ensure_ascii=False)
                else:
                    flat_result[key] = value
            flat_results.append(flat_result)
        
        df = pd.DataFrame(flat_results)
        
        # Создаем Excel файл в памяти
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Результаты поиска', index=False)
            
            # Получаем лист
            worksheet = writer.sheets['Результаты поиска']
            
            # Форматирование заголовков
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
            
            # Автоширина колонок
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        cell_str = str(cell.value)
                    except Exception:
                        # Игнорируем ошибки преобразования значения ячейки в строку при расчете ширины колонки
                        continue

                    if len(cell_str) > max_length:
                        max_length = len(cell_str)
                
                adjusted_width = min(max_length + 2, 50)  # Максимум 50 символов
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Цветовое кодирование по релевантности (если есть колонка relevance)
            if 'relevance' in df.columns:
                relevance_col_idx = list(df.columns).index('relevance') + 1
                
                for row_idx in range(2, len(df) + 2):  # Начинаем со 2-й строки (после заголовка)
                    cell = worksheet.cell(row=row_idx, column=relevance_col_idx)
                    try:
                        relevance = float(cell.value)
                        
                        # Зеленый для высокой релевантности
                        if relevance >= 0.8:
                            cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                        # Желтый для средней
                        elif relevance >= 0.5:
                            cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                        # Красный для низкой
                        else:
                            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    except Exception:
                        # Игнорируем ошибки при вычислении релевантности для цветового кодирования
                        pass
            
            # Добавляем фильтры
            worksheet.auto_filter.ref = worksheet.dimensions
        
        output.seek(0)
        return output
    
    @staticmethod
    def create_batch_excel(batch_results: Dict[str, List[Dict]], filename: str = None):
        """
        Создать Excel файл с несколькими листами для массового экспорта
        
        Args:
            batch_results: Словарь {query: results}
        """
        try:
            import pandas as pd
        except ImportError:
            raise ImportError("Для экспорта в Excel требуется pandas")
        
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for query, results in batch_results.items():
                if not results:
                    continue
                
                # Убираем вложенные объекты
                flat_results = []
                for result in results:
                    flat_result = {}
                    for key, value in result.items():
                        if isinstance(value, (dict, list)):
                            flat_result[key] = json.dumps(value, ensure_ascii=False)
                        else:
                            flat_result[key] = value
                    flat_results.append(flat_result)
                
                df = pd.DataFrame(flat_results)
                
                # Имя листа (первые 31 символ, т.к. Excel ограничивает длину)
                sheet_name = query[:31] if len(query) > 31 else query
                # Заменяем недопустимые символы
                sheet_name = sheet_name.replace('/', '_').replace('\\', '_').replace(':', '_')
                
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Форматирование
                worksheet = writer.sheets[sheet_name]
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            cell_value = str(cell.value)
                        except Exception:
                            # Если значение ячейки не приводится к строке, пропускаем его при расчёте ширины столбца
                            continue
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        return output


class AnalogsExporter:
    """Экспорт аналогов подшипников"""
    
    @staticmethod
    def to_json(analogs: List[Dict], pretty: bool = True) -> str:
        """Экспорт аналогов в JSON"""
        return SearchResultsExporter.to_json(analogs, pretty)
    
    @staticmethod
    def to_csv(analogs: List[Dict]) -> str:
        """Экспорт аналогов в CSV"""
        if not analogs:
            return ""
        
        output = io.StringIO()
        
        # Определяем поля для аналогов
        fieldnames = ['code', 'standard', 'manufacturer', 'compatibility', 'notes']
        
        writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        
        for analog in analogs:
            writer.writerow(analog)
        
        return output.getvalue()
    
    @staticmethod
    def to_excel(analogs: List[Dict], filename: str = None):
        """Экспорт аналогов в Excel"""
        try:
            import pandas as pd
        except ImportError:
            raise ImportError("Для экспорта в Excel требуется pandas")
        
        if not analogs:
            return None
        
        df = pd.DataFrame(analogs)
        
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Аналоги', index=False)
            
            worksheet = writer.sheets['Аналоги']
            
            # Автоширина колонок
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        cell_str = str(cell.value)
                    except Exception:
                        # Игнорируем ошибки при вычислении длины значения ячейки
                        continue
                    
                    if len(cell_str) > max_length:
                        max_length = len(cell_str)
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        return output
